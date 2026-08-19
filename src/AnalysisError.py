import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
# ============================================================================
# 1. THRESHOLD SENSITIVITY ANALYSIS
# ============================================================================

def extract_triple_scores(data):
    """Extract all per-triple scores for threshold analysis"""
    scores = []
    for entry in data:
        if not isinstance(entry, list):
            continue
        for sentence_group in entry:
            if not isinstance(sentence_group, list):
                continue
            for sentence_data in sentence_group:
                if not isinstance(sentence_data, dict):
                    continue
                if 'per_triple_scores' in sentence_data:
                    for triple_score in sentence_data['per_triple_scores']:
                        scores.append({
                            'relevance': triple_score.get('relevance', 0),
                            'consistency': triple_score.get('consistency', 0),
                            'combined': triple_score.get('combined', 0),
                            'z_score': triple_score.get('z_score', 0),
                            'kept': triple_score.get('kept', False)
                        })
    return pd.DataFrame(scores)

def threshold_sensitivity_analysis(scores_df):
    """Analyze how different thresholds affect precision and recall"""
    results = []
    
    # Test different combined score thresholds (0.3 to 0.9)
    for threshold in np.arange(0.3, 0.91, 0.05):
        predicted_kept = scores_df['combined'] >= threshold
        true_kept = scores_df['kept']
        
        tp = ((predicted_kept) & (true_kept)).sum()
        fp = ((predicted_kept) & (~true_kept)).sum()
        fn = ((~predicted_kept) & (true_kept)).sum()
        
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
        
        results.append({
            'threshold': threshold,
            'precision': precision,
            'recall': recall,
            'f1': f1,
            'tp': tp,
            'fp': fp,
            'fn': fn
        })
    
    # Test different z-score thresholds
    for threshold in np.arange(-2.0, 0.51, 0.25):
        # Z-score: keep if z_score > threshold (higher z-score = more confident)
        predicted_kept = scores_df['z_score'] >= threshold
        true_kept = scores_df['kept']
        
        tp = ((predicted_kept) & (true_kept)).sum()
        fp = ((predicted_kept) & (~true_kept)).sum()
        fn = ((~predicted_kept) & (true_kept)).sum()
        
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
        
        results.append({
            'threshold_type': 'z_score',
            'threshold': threshold,
            'precision': precision,
            'recall': recall,
            'f1': f1,
            'tp': tp,
            'fp': fp,
            'fn': fn
        })
    
    return pd.DataFrame(results)

# ============================================================================
# 2. PRECISION PERFORMANCE TABLE
# ============================================================================

def extract_sentence_statistics(data):
    """Extract statistics per sentence"""
    stats = []
    for entry in data:
        if not isinstance(entry, list):
            continue
        for sentence_group in entry:
            if not isinstance(sentence_group, list):
                continue
            for sentence_data in sentence_group:
                if not isinstance(sentence_data, dict):
                    continue
                if 'total_triples' in sentence_data and 'kept_triples' in sentence_data:
                    stats.append({
                        'total_triples': sentence_data['total_triples'],
                        'kept_triples': sentence_data['kept_triples'],
                        'filtered_out': sentence_data['filtered_out'],
                        'precision': sentence_data['precision'],
                        'error_breakdown': sentence_data.get('error_breakdown', {})
                    })
    return pd.DataFrame(stats)

# ============================================================================
# 3. CATEGORICAL ERROR DISTRIBUTION TABLE
# ============================================================================

def extract_error_distribution(data):
    """Extract error distribution across sentence categories"""
    error_counts = defaultdict(int)
    total_sentences = 0
    
    for entry in data:
        if not isinstance(entry, list):
            continue
        for sentence_group in entry:
            if not isinstance(sentence_group, list):
                continue
            for sentence_data in sentence_group:
                if not isinstance(sentence_data, dict):
                    continue
                total_sentences += 1
                if 'error_breakdown' in sentence_data:
                    for error_type, count in sentence_data['error_breakdown'].items():
                        error_counts[error_type] += count
    
    return dict(error_counts), total_sentences

# ============================================================================
# 4. DETAILED PERFORMANCE METRICS TABLE
# ============================================================================

def extract_performance_metrics(data):
    """Extract detailed performance metrics"""
    metrics = []
    for entry in data:
        if not isinstance(entry, list):
            continue
        for sentence_group in entry:
            if not isinstance(sentence_group, list):
                continue
            for idx, sentence_data in enumerate(sentence_group):
                if not isinstance(sentence_data, dict):
                    continue
                # Try to get the sentence text from context
                sentence_text = "N/A"
                if idx > 0 and isinstance(sentence_group[idx-1], dict):
                    sentence_text = sentence_group[idx-1].get('sentence', 'N/A')
                elif isinstance(sentence_group, list) and len(sentence_group) > 0:
                    # Check if first element is the sentence text
                    if idx == 0 and isinstance(sentence_group[0], str):
                        sentence_text = sentence_group[0]
                
                if 'per_triple_scores' in sentence_data:
                    scores = sentence_data['per_triple_scores']
                    avg_relevance = np.mean([s.get('relevance', 0) for s in scores])
                    avg_consistency = np.mean([s.get('consistency', 0) for s in scores])
                    avg_combined = np.mean([s.get('combined', 0) for s in scores])
                    kept_count = sum(1 for s in scores if s.get('kept', False))
                    
                    metrics.append({
                        'sentence': sentence_text[:50] + '...' if len(sentence_text) > 50 else sentence_text,
                        'total_triples': sentence_data.get('total_triples', 0),
                        'kept_triples': sentence_data.get('kept_triples', 0),
                        'filtered_out': sentence_data.get('filtered_out', 0),
                        'precision': sentence_data.get('precision', 0),
                        'avg_relevance': avg_relevance,
                        'avg_consistency': avg_consistency,
                        'avg_combined': avg_combined,
                        'error_type': list(sentence_data.get('error_breakdown', {}).keys())[0] if sentence_data.get('error_breakdown', {}) else 'none'
                    })
    return pd.DataFrame(metrics)

# ============================================================================
# 5. ERROR ANALYSIS - SPECIFIC FAILURE CASES
# ============================================================================

def identify_failure_cases(data):
    """Identify and analyze specific failure cases"""
    failures = []
    
    for entry in data:
        if not isinstance(entry, list):
            continue
        for sentence_group in entry:
            if not isinstance(sentence_group, list):
                continue
            for idx, sentence_data in enumerate(sentence_group):
                if not isinstance(sentence_data, dict):
                    continue
                if 'per_triple_scores' in sentence_data:
                    for triple_score in sentence_data['per_triple_scores']:
                        if not triple_score.get('kept', False):
                            error_type = triple_score.get('error_type', 'unknown')
                            failures.append({
                                'error_type': error_type,
                                'relevance': triple_score.get('relevance', 0),
                                'consistency': triple_score.get('consistency', 0),
                                'combined': triple_score.get('combined', 0),
                                'z_score': triple_score.get('z_score', 0),
                                'triple': triple_score.get('triple', [])
                            })
    return pd.DataFrame(failures)
# ============================================================================
# 6. VISUALIZATION OF THRESHOLD SENSITIVITY
# ============================================================================

def plot_threshold_sensitivity(results_df,combined_optimal,zscore_optimal,args):
    """Create visualization of threshold sensitivity"""
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))
    
    # Combined score threshold plot
    combined_data = results_df[results_df['threshold_type'].isna()]
    ax1 = axes[0]
    ax1.plot(combined_data['threshold'], combined_data['precision'], 'b-', label='Precision', linewidth=2)
    ax1.plot(combined_data['threshold'], combined_data['recall'], 'r-', label='Recall', linewidth=2)
    ax1.plot(combined_data['threshold'], combined_data['f1'], 'g-', label='F1', linewidth=2)
    ax1.axvline(x=combined_optimal['threshold'], color='k', linestyle='--', alpha=0.5, label=f"Optimal: {combined_optimal['threshold']:.2f}")
    ax1.set_xlabel('Combined Score Threshold', fontsize=12)
    ax1.set_ylabel('Score', fontsize=12)
    ax1.set_title('Combined Score Threshold Sensitivity', fontsize=14)
    ax1.legend()
    ax1.grid(True, alpha=0.3)
    
    # Z-score threshold plot
    zscore_data = results_df[results_df['threshold_type'] == 'z_score']
    ax2 = axes[1]
    ax2.plot(zscore_data['threshold'], zscore_data['precision'], 'b-', label='Precision', linewidth=2)
    ax2.plot(zscore_data['threshold'], zscore_data['recall'], 'r-', label='Recall', linewidth=2)
    ax2.plot(zscore_data['threshold'], zscore_data['f1'], 'g-', label='F1', linewidth=2)
    ax2.axvline(x=zscore_optimal['threshold'], color='k', linestyle='--', alpha=0.5, label=f"Optimal: {zscore_optimal['threshold']:.2f}")
    ax2.set_xlabel('Z-Score Threshold', fontsize=12)
    ax2.set_ylabel('Score', fontsize=12)
    ax2.set_title('Z-Score Threshold Sensitivity', fontsize=14)
    ax2.legend()
    ax2.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(args.PathDataset+'threshold_sensitivity.png', dpi=300, bbox_inches='tight')
    #plt.show()
    plt.close(fig)
# ============================================================================
# 8. TOTAL RELATIONS COUNT
# ============================================================================

def count_unique_relations(data):
    """Count unique relations across all triples"""
    relations = set()
    for entry in data:
        if not isinstance(entry, list):
            continue
        for sentence_group in entry:
            if not isinstance(sentence_group, list):
                continue
            for sentence_data in sentence_group:
                if not isinstance(sentence_data, dict):
                    continue
                if 'per_triple_scores' in sentence_data:
                    for triple_score in sentence_data['per_triple_scores']:
                        triple = triple_score.get('triple', [])
                        if len(triple) >= 2:
                            relations.add(triple[1])  # The relation/predicate
    
    return list(relations)
# ============================================================================
# 9. ERROR PATTERN ANALYSIS
# ============================================================================

def analyze_error_patterns(failure_cases):
    """Analyze patterns in failed triples"""
    print("\n" + "=" * 80)
    print("ERROR PATTERN ANALYSIS")
    print("=" * 80)
    
    if len(failure_cases) == 0:
        print("No failures found.")
        return
    
    # Analyze by error type
    print("\nError Type Distribution:")
    for error_type, group in failure_cases.groupby('error_type'):
        print(f"\n{error_type}:")
        print(f"  Count: {len(group)}")
        print(f"  Avg Relevance: {group['relevance'].mean():.3f}")
        print(f"  Avg Consistency: {group['consistency'].mean():.3f}")
        print(f"  Avg Combined: {group['combined'].mean():.3f}")
        print(f"  Avg Z-score: {group['z_score'].mean():.3f}")
    
    # Common patterns in failed triples
    print("\n" + "-" * 40)
    print("COMMON PATTERNS IN FAILED TRIPLES:")
    print("-" * 40)
    
    # Analyze relation patterns
    relation_patterns = failure_cases['triple'].apply(lambda x: x[1] if len(x) > 1 else 'unknown')
    print("\nMost common relations in failed triples:")
    for relation, count in relation_patterns.value_counts().head(10).items():
        print(f"  {relation}: {count}")
    
    # Analyze subject patterns
    subject_patterns = failure_cases['triple'].apply(lambda x: x[0] if len(x) > 0 else 'unknown')
    print("\nMost common subjects in failed triples:")
    for subject, count in subject_patterns.value_counts().head(10).items():
        print(f"  {subject}: {count}")
# ============================================================================
# 10. GENERATE ALL TABLES IN MARKDOWN FORMAT
# ============================================================================

def generate_markdown_tables(precision_table,error_df,metrics_df,combined_optimal,zscore_optimal,failure_cases):
    """Generate all tables in markdown format for easy inclusion in reports"""
    
    md_output = []
    
    # Table 1: Precision Performance
    md_output.append("# TABLES FOR ADAPTIVETRIPLEFILTER ANALYSIS\n")
    md_output.append("## Table 1: Precision Performance by Sentence Type\n")
    md_output.append("| Total Triples | Mean Precision | Std Precision | Min Precision | Max Precision | Avg Kept | Avg Filtered |")
    md_output.append("|--------------|----------------|---------------|---------------|---------------|----------|--------------|")
    
    for idx, row in precision_table.iterrows():
        md_output.append(f"| {idx} | {row[('precision', 'mean')]:.3f} | {row[('precision', 'std')]:.3f} | {row[('precision', 'min')]:.3f} | {row[('precision', 'max')]:.3f} | {row[('kept_triples', 'mean')]:.1f} | {row[('filtered_out', 'mean')]:.1f} |")
    
    # Table 2: Error Distribution
    md_output.append("\n## Table 2: Categorical Error Distribution\n")
    md_output.append("| Error Type | Count | Percentage |")
    md_output.append("|------------|-------|------------|")
    
    for _, row in error_df.iterrows():
        md_output.append(f"| {row['Error Type']} | {row['Count']} | {row['Percentage']:.1f}% |")
    
    # Table 3: Performance Metrics
    md_output.append("\n## Table 3: Performance Metrics and Error Distribution\n")
    md_output.append("| Sentence | Triples | Kept | Filtered | Precision | Avg Relevance | Avg Consistency | Error Type |")
    md_output.append("|----------|---------|------|----------|-----------|---------------|-----------------|------------|")
    
    for _, row in metrics_df.head(10).iterrows():
        md_output.append(f"| {row['sentence'][:30]}... | {row['total_triples']} | {row['kept_triples']} | {row['filtered_out']} | {row['precision']:.3f} | {row['avg_relevance']:.3f} | {row['avg_consistency']:.3f} | {row['error_type']} |")
    
    # Table 4: Optimal Thresholds
    md_output.append("\n## Table 4: Optimal Threshold Settings\n")
    md_output.append("| Threshold Type | Optimal Value | Precision | Recall | F1-Score |")
    md_output.append("|----------------|---------------|-----------|--------|----------|")
    md_output.append(f"| Combined Score | {combined_optimal['threshold']:.2f} | {combined_optimal['precision']:.3f} | {combined_optimal['recall']:.3f} | {combined_optimal['f1']:.3f} |")
    md_output.append(f"| Z-Score | {zscore_optimal['threshold']:.2f} | {zscore_optimal['precision']:.3f} | {zscore_optimal['recall']:.3f} | {zscore_optimal['f1']:.3f} |")
    
    # Table 5: Failure Case Summary
    md_output.append("\n## Table 5: Failure Case Summary\n")
    md_output.append("| Error Type | Count | Avg Relevance | Avg Consistency | Avg Z-Score |")
    md_output.append("|------------|-------|---------------|-----------------|-------------|")
    try:
        for error_type, group in failure_cases.groupby('error_type'):
            md_output.append(f"| {error_type} | {len(group)} | {group['relevance'].mean():.3f} | {group['consistency'].mean():.3f} | {group['z_score'].mean():.3f} |")
    except:
        pass
    return "\n".join(md_output)
def save_tables_to_excel(precision_table, error_df, metrics_df, combined_optimal, zscore_optimal, failure_cases, filename='analysis_results.xlsx'):
    """Save all tables to an Excel file with proper formatting"""
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        
        # Table 1: Precision Performance
        precision_table.to_excel(writer, sheet_name='Precision Performance')
        
        # Table 2: Error Distribution
        error_df.to_excel(writer, sheet_name='Error Distribution', index=False)
        
        # Table 3: Performance Metrics (all rows, not just top 10)
        metrics_df.to_excel(writer, sheet_name='Performance Metrics', index=False)
        
        # Table 4: Optimal Thresholds
        thresholds_df = pd.DataFrame({
            'Threshold Type': ['Combined Score', 'Z-Score'],
            'Optimal Value': [combined_optimal['threshold'], zscore_optimal['threshold']],
            'Precision': [combined_optimal['precision'], zscore_optimal['precision']],
            'Recall': [combined_optimal['recall'], zscore_optimal['recall']],
            'F1-Score': [combined_optimal['f1'], zscore_optimal['f1']]
        })
        thresholds_df.to_excel(writer, sheet_name='Optimal Thresholds', index=False)
        
        # Table 5: Failure Case Summary
        try:
            failure_summary = failure_cases.groupby('error_type').agg({
                'relevance': 'mean',
                'consistency': 'mean',
                'z_score': 'mean'
            }).reset_index()
            failure_summary.columns = ['Error Type', 'Avg Relevance', 'Avg Consistency', 'Avg Z-Score']
            failure_summary.to_excel(writer, sheet_name='Failure Summary', index=False)
        except:
            pass
        # Format the Excel file
        format_excel_sheets(writer)

def format_excel_sheets(writer):
    """Apply formatting to all sheets in the Excel file"""
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # Auto-adjust column widths
        for col_idx, column in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
def main_analysis(data,args):
    scores_df = extract_triple_scores(data)
    threshold_results = threshold_sensitivity_analysis(scores_df)

    print("=" * 80)
    print("THRESHOLD SENSITIVITY ANALYSIS")
    print("=" * 80)
    print("\nCombined Score Threshold Analysis:")
    print(threshold_results[threshold_results['threshold_type'].isna()].to_string(index=False))

    print("\nZ-Score Threshold Analysis:")
    print(threshold_results[threshold_results['threshold_type'] == 'z_score'].to_string(index=False))

    # Find optimal thresholds
    combined_optimal = threshold_results[threshold_results['threshold_type'].isna()].loc[
        threshold_results[threshold_results['threshold_type'].isna()]['f1'].idxmax()
    ]
    zscore_optimal = threshold_results[threshold_results['threshold_type'] == 'z_score'].loc[
        threshold_results[threshold_results['threshold_type'] == 'z_score']['f1'].idxmax()
    ]

    print("\n" + "=" * 80)
    print("OPTIMAL THRESHOLDS")
    print("=" * 80)
    print(f"Optimal Combined Score Threshold: {combined_optimal['threshold']:.2f}")
    print(f"  - Precision: {combined_optimal['precision']:.3f}")
    print(f"  - Recall: {combined_optimal['recall']:.3f}")
    print(f"  - F1: {combined_optimal['f1']:.3f}")
    print(f"\nOptimal Z-Score Threshold: {zscore_optimal['threshold']:.2f}")
    print(f"  - Precision: {zscore_optimal['precision']:.3f}")
    print(f"  - Recall: {zscore_optimal['recall']:.3f}")
    print(f"  - F1: {zscore_optimal['f1']:.3f}")
    # Extract sentence statistics
    sentence_stats = extract_sentence_statistics(data)

    # Precision Performance Table
    print("\n" + "=" * 80)
    print("TABLE 1: PRECISION PERFORMANCE")
    print("=" * 80)

    precision_table = sentence_stats.groupby('total_triples').agg({
        'precision': ['mean', 'std', 'min', 'max'],
        'kept_triples': 'mean',
        'filtered_out': 'mean'
    }).round(3)

    print(precision_table.to_string())


    error_counts, total_sentences = extract_error_distribution(data)

    print("\n" + "=" * 80)
    print("TABLE 2: CATEGORICAL ERROR DISTRIBUTION")
    print("=" * 80)
    if len(error_counts)>0:
        error_df = pd.DataFrame([
            {
                'Error Type': error_type,
                'Count': count,
                'Percentage': (count / total_sentences) * 100
            }
            for error_type, count in error_counts.items()
        ]).sort_values('Count', ascending=False)

        print(error_df.to_string(index=False))
    else:
        error_df=pd.DataFrame([
            {
                'Error Type': '',
                'Count': 0,
                'Percentage': (0 / 1) * 100
            }])     



    metrics_df = extract_performance_metrics(data)

    print("\n" + "=" * 80)
    print("TABLE 3: PERFORMANCE METRICS BY SENTENCE")
    print("=" * 80)
    print(metrics_df.head(10).to_string(index=False))



    failure_cases = identify_failure_cases(data)

    print("\n" + "=" * 80)
    print("FAILURE CASE ANALYSIS")
    print("=" * 80)

    if len(failure_cases) > 0:
        print(f"\nTotal Failed Triples: {len(failure_cases)}")
        print("\nFailure Type Distribution:")
        print(failure_cases['error_type'].value_counts().to_string())
    
        print("\n" + "-" * 40)
        print("EXAMPLE FAILURE CASES:")
        print("-" * 40)
    
        for error_type in failure_cases['error_type'].unique():
            examples = failure_cases[failure_cases['error_type'] == error_type].head(2)
            print(f"\nError Type: {error_type}")
            for _, row in examples.iterrows():
                triple = row['triple']
                print(f"  Triple: ({triple[0]}, {triple[1]}, {triple[2]})")
                print(f"  Relevance: {row['relevance']:.3f}, Consistency: {row['consistency']:.3f}")
                print(f"  Combined: {row['combined']:.3f}, Z-score: {row['z_score']:.3f}")
            print()

    plot_threshold_sensitivity(threshold_results,combined_optimal,zscore_optimal,args)

    # ============================================================================
    # 7. SUMMARY STATISTICS
    # ============================================================================

    print("\n" + "=" * 80)
    print("SUMMARY STATISTICS")
    print("=" * 80)

    total_triples = sentence_stats['total_triples'].sum()
    kept_triples = sentence_stats['kept_triples'].sum()
    filtered_out = sentence_stats['filtered_out'].sum()
    avg_precision = sentence_stats['precision'].mean()
    weighted_precision = kept_triples / total_triples if total_triples > 0 else 0

    print(f"Total Triples Extracted: {total_triples}")
    print(f"Triples Kept: {kept_triples}")
    print(f"Triples Filtered Out: {filtered_out}")
    print(f"Filter Rate: {filtered_out/total_triples*100:.2f}%")
    print(f"Average Precision per Sentence: {avg_precision:.3f}")
    print(f"Weighted Precision: {weighted_precision:.3f}")



    unique_relations = count_unique_relations(data)
    print(f"\nTotal Unique Relations Identified: {len(unique_relations)}")



    analyze_error_patterns(failure_cases)
    

    save_tables_to_excel(precision_table, error_df, metrics_df, combined_optimal, zscore_optimal, failure_cases, filename=args.PathDataset + 'analysis_results.xlsx')

    # Write tables to file
    with open(args.PathDataset+'analysis_tables.md', 'w') as f:
        f.write(generate_markdown_tables(precision_table,error_df,metrics_df,combined_optimal,zscore_optimal,failure_cases))

    print("\n" + "=" * 80)
    print("TABLES GENERATED")
    print("=" * 80)
    print("Tables have been written to 'analysis_tables.md'")
    print("Threshold sensitivity plot saved to 'threshold_sensitivity.png'")

    # ============================================================================
    # 11. RECOMMENDATIONS FOR IMPROVEMENT
    # ============================================================================

    print("\n" + "=" * 80)
    print("RECOMMENDATIONS FOR IMPROVEMENT")
    print("=" * 80)

    print("""
    1. THRESHOLD OPTIMIZATION:
       - The optimal combined score threshold is {:.2f}, but consider dynamic thresholds
       - Implement adaptive thresholds based on sentence complexity and domain
       - Consider using ensemble methods with multiple threshold strategies

    2. ERROR PATTERN MITIGATION:
       - For 'outlier_inconsistent' errors: Improve consistency scoring with context-aware embeddings
       - For 'low_context_relevance' errors: Enhance context extraction with cross-sentence attention
       - Consider specialized handling for technical domain abstracts

    3. RELATION DIVERSITY:
       - Total unique relations: {}
       - Consider expanding relation vocabulary with domain-specific terminology
       - Implement relation pattern learning from successful extractions

    4. FUTURE WORK:
       - Active learning to adapt thresholds for challenging sentence patterns
       - Multi-stage filtering with progressive threshold adjustment
       - Integration with knowledge graphs for validation of extracted relations
    """.format(combined_optimal['threshold'], len(unique_relations)))


